"""Format-specific data schema summarizers."""

from __future__ import annotations

import csv
import json
import sqlite3
from itertools import islice
from pathlib import Path

from .._textutils import _json_dumps
from ..constants import TEXT_ENCODINGS
from ..models import DataSummary, DataSummarySection, ExtractionFailure
from .builders import (
    DATA_SCHEMA_EXTRACTOR_NAME,
    DATA_SCHEMA_EXTRACTOR_VERSION,
    JSON_PREVIEW_CHAR_LIMIT,
    MAX_JSON_FULL_LOAD_BYTES,
    SAMPLE_ROW_LIMIT,
    _build_data_summary,
    _build_section,
    _clean_column_name,
    _deduplicate_column_names,
    _json_safe,
    _merge_fieldnames,
    _read_text_preview,
    _row_from_sequence,
)


def summarize_csv(path: Path, file_id: int = 0) -> DataSummary:
    for encoding in TEXT_ENCODINGS:
        try:
            return _summarize_csv_with_encoding(path, file_id, encoding)
        except UnicodeDecodeError:
            continue
    raise ExtractionFailure(
        "could not decode CSV with supported text encodings",
        extractor_name=DATA_SCHEMA_EXTRACTOR_NAME,
        extractor_version=DATA_SCHEMA_EXTRACTOR_VERSION,
        metadata_json=_json_dumps({"extension": ".csv"}),
    )


def summarize_xlsx(path: Path, file_id: int = 0) -> DataSummary:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExtractionFailure(
            "openpyxl is required to summarize XLSX files",
            extractor_name=DATA_SCHEMA_EXTRACTOR_NAME,
            extractor_version=DATA_SCHEMA_EXTRACTOR_VERSION,
            metadata_json=_json_dumps({"extension": ".xlsx"}),
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sections: list[DataSummarySection] = []
        for worksheet in workbook.worksheets:
            row_iter = worksheet.iter_rows(values_only=True)
            header_row = next(row_iter, None)
            if header_row is None:
                headers: list[str] = []
                sample_rows: list[dict[str, object]] = []
                row_count = 0
            else:
                headers = _deduplicate_column_names(
                    [
                        _clean_column_name(value, index)
                        for index, value in enumerate(header_row, start=1)
                    ]
                )
                sample_rows = [
                    _row_from_sequence(headers, row)
                    for row in islice(row_iter, SAMPLE_ROW_LIMIT)
                ]
                row_count = max(int(worksheet.max_row or 0) - 1, 0)
            sections.append(
                _build_section(
                    name=worksheet.title,
                    kind="sheet",
                    location_value=worksheet.title,
                    row_count=row_count,
                    fieldnames=headers,
                    sample_rows=sample_rows,
                )
            )
    finally:
        workbook.close()

    return _build_data_summary(
        file_id=file_id,
        file_format="xlsx",
        sections=sections,
        sheet_count=len(sections),
    )


def summarize_json(path: Path, file_id: int = 0) -> DataSummary:
    if path.stat().st_size > MAX_JSON_FULL_LOAD_BYTES:
        return _summarize_large_json_preview(path, file_id)

    payload = _load_json_with_fallback(path)
    rows, fieldnames, row_count, name = _rows_from_json_payload(payload)
    section = _build_section(
        name=name,
        kind="file",
        location_value="file",
        row_count=row_count,
        fieldnames=fieldnames,
        sample_rows=rows[:SAMPLE_ROW_LIMIT],
    )
    return _build_data_summary(
        file_id=file_id,
        file_format="json",
        sections=[section],
    )


def summarize_jsonl(path: Path, file_id: int = 0) -> DataSummary:
    for encoding in TEXT_ENCODINGS:
        try:
            return _summarize_jsonl_with_encoding(path, file_id, encoding)
        except UnicodeDecodeError:
            continue
    raise ExtractionFailure(
        "could not decode JSONL with supported text encodings",
        extractor_name=DATA_SCHEMA_EXTRACTOR_NAME,
        extractor_version=DATA_SCHEMA_EXTRACTOR_VERSION,
        metadata_json=_json_dumps({"extension": ".jsonl"}),
    )


def summarize_sqlite(path: Path, file_id: int = 0) -> DataSummary:
    uri = f"{path.resolve().as_uri()}?mode=ro"
    try:
        connection = sqlite3.connect(uri, uri=True)
    except sqlite3.Error as exc:
        raise ExtractionFailure(
            f"could not open SQLite database read-only: {exc}",
            extractor_name=DATA_SCHEMA_EXTRACTOR_NAME,
            extractor_version=DATA_SCHEMA_EXTRACTOR_VERSION,
            metadata_json=_json_dumps({"extension": path.suffix.lower()}),
        ) from exc

    connection.row_factory = sqlite3.Row
    try:
        connection.execute("PRAGMA query_only = ON")
        table_rows = connection.execute(
            """
            SELECT name
            FROM sqlite_master
            WHERE type = 'table'
              AND name NOT LIKE 'sqlite_%'
            ORDER BY name
            """
        ).fetchall()
        sections = []
        for table_row in table_rows:
            table_name = str(table_row["name"])
            quoted_name = _quote_identifier(table_name)
            pragma_rows = connection.execute(
                f"PRAGMA table_info({quoted_name})"
            ).fetchall()
            columns = [str(row["name"]) for row in pragma_rows]
            row_count = _sqlite_count_rows(connection, quoted_name)
            sample_rows = [
                dict(row)
                for row in connection.execute(
                    f"SELECT * FROM {quoted_name} LIMIT {SAMPLE_ROW_LIMIT}"
                ).fetchall()
            ]
            section = _build_section(
                name=table_name,
                kind="table",
                location_value=table_name,
                row_count=row_count,
                fieldnames=columns,
                sample_rows=sample_rows,
                column_overrides={
                    str(row["name"]): {
                        "declared_type": str(row["type"]),
                        "not_null": bool(row["notnull"]),
                        "primary_key": bool(row["pk"]),
                    }
                    for row in pragma_rows
                },
            )
            sections.append(section)
    except sqlite3.Error as exc:
        raise ExtractionFailure(
            f"could not inspect SQLite database safely: {exc}",
            extractor_name=DATA_SCHEMA_EXTRACTOR_NAME,
            extractor_version=DATA_SCHEMA_EXTRACTOR_VERSION,
            metadata_json=_json_dumps({"extension": path.suffix.lower()}),
        ) from exc
    finally:
        connection.close()

    return _build_data_summary(
        file_id=file_id,
        file_format="sqlite",
        sections=sections,
        table_count=len(sections),
    )


def _summarize_csv_with_encoding(
    path: Path,
    file_id: int,
    encoding: str,
) -> DataSummary:
    with path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.DictReader(handle)
        raw_fieldnames = reader.fieldnames or ()
        fieldnames = _deduplicate_column_names(
            [
                _clean_column_name(name, index)
                for index, name in enumerate(raw_fieldnames, start=1)
            ]
        )
        sample_rows: list[dict[str, object]] = []
        row_count = 0
        for row in reader:
            row_count += 1
            if len(sample_rows) < SAMPLE_ROW_LIMIT:
                sample_rows.append(
                    {name: _json_safe(row.get(name)) for name in fieldnames}
                )

    section = _build_section(
        name="file",
        kind="file",
        location_value="file",
        row_count=row_count,
        fieldnames=fieldnames,
        sample_rows=sample_rows,
    )
    return _build_data_summary(
        file_id=file_id,
        file_format="csv",
        sections=[section],
    )


def _summarize_jsonl_with_encoding(
    path: Path,
    file_id: int,
    encoding: str,
) -> DataSummary:
    sample_rows: list[dict[str, object]] = []
    row_count = 0
    fieldnames: list[str] = []
    with path.open("r", encoding=encoding) as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ExtractionFailure(
                    f"invalid JSONL on line {line_number}: {exc.msg}",
                    extractor_name=DATA_SCHEMA_EXTRACTOR_NAME,
                    extractor_version=DATA_SCHEMA_EXTRACTOR_VERSION,
                    metadata_json=_json_dumps({"extension": ".jsonl"}),
                ) from exc
            row = _row_from_json_record(payload)
            row_count += 1
            if len(sample_rows) < SAMPLE_ROW_LIMIT:
                sample_rows.append(row)
                fieldnames = _merge_fieldnames(fieldnames, row.keys())

    section = _build_section(
        name="file",
        kind="file",
        location_value="file",
        row_count=row_count,
        fieldnames=fieldnames,
        sample_rows=sample_rows,
    )
    return _build_data_summary(
        file_id=file_id,
        file_format="jsonl",
        sections=[section],
    )


def _load_json_with_fallback(path: Path) -> object:
    last_decode_error: UnicodeDecodeError | None = None
    for encoding in TEXT_ENCODINGS:
        try:
            with path.open("r", encoding=encoding) as handle:
                return json.load(handle)
        except UnicodeDecodeError as exc:
            last_decode_error = exc
            continue
        except json.JSONDecodeError as exc:
            raise ExtractionFailure(
                f"invalid JSON file: {exc.msg}",
                extractor_name=DATA_SCHEMA_EXTRACTOR_NAME,
                extractor_version=DATA_SCHEMA_EXTRACTOR_VERSION,
                metadata_json=_json_dumps({"extension": ".json"}),
            ) from exc
    raise ExtractionFailure(
        "could not decode JSON with supported text encodings",
        extractor_name=DATA_SCHEMA_EXTRACTOR_NAME,
        extractor_version=DATA_SCHEMA_EXTRACTOR_VERSION,
        metadata_json=_json_dumps({"extension": ".json"}),
    ) from last_decode_error


def _summarize_large_json_preview(path: Path, file_id: int) -> DataSummary:
    preview = _read_text_preview(path, JSON_PREVIEW_CHAR_LIMIT)
    section = _build_section(
        name="file",
        kind="file",
        location_value="file",
        row_count=None,
        fieldnames=("preview",),
        sample_rows=[{"preview": preview}],
    )
    return _build_data_summary(
        file_id=file_id,
        file_format="json",
        sections=[section],
        metadata={"large_json_preview": True},
    )


def _rows_from_json_payload(
    payload: object,
) -> tuple[list[dict[str, object]], list[str], int, str]:
    if isinstance(payload, list):
        rows = [_row_from_json_record(item) for item in payload[:SAMPLE_ROW_LIMIT]]
        fieldnames: list[str] = []
        for row in rows:
            fieldnames = _merge_fieldnames(fieldnames, row.keys())
        return rows, fieldnames, len(payload), "array"
    if isinstance(payload, dict):
        row = {str(key): _json_safe(value) for key, value in payload.items()}
        return [row], list(row.keys()), 1, "object"
    row = {"value": _json_safe(payload)}
    return [row], ["value"], 1, "value"


def _row_from_json_record(payload: object) -> dict[str, object]:
    if isinstance(payload, dict):
        return {str(key): _json_safe(value) for key, value in payload.items()}
    return {"value": _json_safe(payload)}


def _sqlite_count_rows(connection: sqlite3.Connection, quoted_table: str) -> int:
    row = connection.execute(f"SELECT COUNT(*) AS count FROM {quoted_table}").fetchone()
    return int(row["count"] if row else 0)


def _quote_identifier(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'
